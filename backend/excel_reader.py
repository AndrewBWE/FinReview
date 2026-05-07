from __future__ import annotations

import time
from io import BytesIO
from typing import Any

import openpyxl


def run_excel(file_bytes: bytes, filename: str) -> dict[str, Any]:
    start = time.time()

    wb = openpyxl.load_workbook(
        filename=BytesIO(file_bytes), read_only=True, data_only=True
    )

    pages: list[dict[str, Any]] = []

    for sheet_index, sheet_name in enumerate(wb.sheetnames):
        ws = wb[sheet_name]

        raw_rows: list[list[str]] = []
        for row in ws.iter_rows(values_only=True):
            cells = [str(cell) if cell is not None else "" for cell in row]
            raw_rows.append(cells)

        # Drop fully empty rows and trim trailing empty columns
        rows = [r for r in raw_rows if any(c.strip() for c in r)]
        if not rows:
            continue
        max_cols = max(len(r) for r in rows)
        rows = [r + [""] * (max_cols - len(r)) for r in rows]

        table_cells = [
            {"row": ri, "col": ci, "content": val, "row_span": 1, "col_span": 1}
            for ri, row in enumerate(rows)
            for ci, val in enumerate(row)
        ]
        table = {
            "row_count": len(rows),
            "column_count": max_cols,
            "cells": table_cells,
        }

        text_lines = [f"Sheet: {sheet_name}", ""]
        for row in rows:
            text_lines.append("\t".join(row))
        text = "\n".join(text_lines)

        pages.append({
            "page_number": sheet_index + 1,
            "text": text,
            "tables": [table],
        })

    wb.close()

    duration_ms = int((time.time() - start) * 1000)
    full_text = "\n\n--- Sheet Break ---\n\n".join(p["text"] for p in pages)

    return {
        "pages": pages,
        "page_count": len(pages),
        "full_text": full_text,
        "duration_ms": duration_ms,
    }
